"""Business logic for financial dashboard and reports."""

import io
import logging
from datetime import date, timedelta
from typing import Optional
from uuid import UUID

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from domains.jusmonitoria.db.models.contrato import Contrato, StatusContrato
from domains.jusmonitoria.db.models.fatura import Fatura, StatusFatura
from domains.jusmonitoria.db.models.lancamento import Lancamento, TipoLancamento
from domains.jusmonitoria.db.repositories.fatura import FaturaRepository
from domains.jusmonitoria.db.repositories.lancamento import LancamentoRepository

logger = logging.getLogger(__name__)


class FinanceiroService:
    """Service for financial metrics and reports."""

    def __init__(self, session: AsyncSession, tenant_id: UUID):
        self.session = session
        self.tenant_id = tenant_id
        self.fatura_repo = FaturaRepository(session, tenant_id)
        self.lancamento_repo = LancamentoRepository(session, tenant_id)

    async def get_dashboard(self, meses: int = 6) -> dict:
        """Get financial dashboard data."""
        today = date.today()
        date_from = date(today.year, today.month, 1) - timedelta(days=meses * 31)

        # Revenue summary from Faturas
        resumo = await self.fatura_repo.get_revenue_summary(date_from, today)

        # Add Lancamento RECEITA to the summary (webhook payments create
        # Lancamentos, not Faturas, so we must include both sources).
        lancamento_receita = await self._get_lancamento_receita_total(date_from, today)
        resumo["total_recebido"] = resumo["total_recebido"] + lancamento_receita

        # Active contracts count
        contratos_query = (
            select(func.count(Contrato.id))
            .where(Contrato.tenant_id == self.tenant_id)
            .where(Contrato.status == StatusContrato.ATIVO)
        )
        contratos_result = await self.session.execute(contratos_query)
        contratos_ativos = contratos_result.scalar_one()

        # Pending invoices count
        pendentes_query = (
            select(func.count(Fatura.id))
            .where(Fatura.tenant_id == self.tenant_id)
            .where(Fatura.status == StatusFatura.PENDENTE)
        )
        pendentes_result = await self.session.execute(pendentes_query)
        faturas_pendentes = pendentes_result.scalar_one()

        # Overdue invoices count
        vencidas_query = (
            select(func.count(Fatura.id))
            .where(Fatura.tenant_id == self.tenant_id)
            .where(Fatura.status.in_([StatusFatura.VENCIDA, StatusFatura.PENDENTE]))
            .where(Fatura.data_vencimento < today)
        )
        vencidas_result = await self.session.execute(vencidas_query)
        faturas_vencidas = vencidas_result.scalar_one()

        # Revenue by month (last N months)
        receita_por_mes = await self._get_receita_por_mes(meses)

        return {
            "resumo": resumo,
            "contratos_ativos": contratos_ativos,
            "faturas_pendentes": faturas_pendentes,
            "faturas_vencidas": faturas_vencidas,
            "receita_por_mes": receita_por_mes,
        }

    async def _get_lancamento_receita_total(self, date_from: date, date_to: date) -> float:
        """Sum of Lancamento RECEITA entries in the period (webhook payments)."""
        query = (
            select(func.coalesce(func.sum(Lancamento.valor), 0))
            .where(Lancamento.tenant_id == self.tenant_id)
            .where(Lancamento.tipo == TipoLancamento.RECEITA)
            .where(Lancamento.data_lancamento >= date_from)
            .where(Lancamento.data_lancamento <= date_to)
        )
        result = await self.session.execute(query)
        return float(result.scalar_one())

    async def _get_receita_por_mes(self, meses: int) -> list[dict]:
        """Get revenue grouped by month."""
        today = date.today()
        result = []

        for i in range(meses - 1, -1, -1):
            # Calculate month boundaries
            month = today.month - i
            year = today.year
            while month <= 0:
                month += 12
                year -= 1

            first_day = date(year, month, 1)
            if month == 12:
                last_day = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                last_day = date(year, month + 1, 1) - timedelta(days=1)

            # Faturado
            faturado_query = (
                select(func.coalesce(func.sum(Fatura.valor), 0))
                .where(Fatura.tenant_id == self.tenant_id)
                .where(Fatura.data_vencimento >= first_day)
                .where(Fatura.data_vencimento <= last_day)
            )
            faturado_result = await self.session.execute(faturado_query)
            valor_faturado = float(faturado_result.scalar_one())

            # Recebido (Faturas pagas)
            recebido_query = (
                select(func.coalesce(func.sum(Fatura.valor_pago), 0))
                .where(Fatura.tenant_id == self.tenant_id)
                .where(Fatura.status == StatusFatura.PAGA)
                .where(Fatura.data_pagamento >= first_day)
                .where(Fatura.data_pagamento <= last_day)
            )
            recebido_result = await self.session.execute(recebido_query)
            valor_recebido_faturas = float(recebido_result.scalar_one())

            # Recebido (Lancamentos RECEITA — webhook payments)
            lanc_query = (
                select(func.coalesce(func.sum(Lancamento.valor), 0))
                .where(Lancamento.tenant_id == self.tenant_id)
                .where(Lancamento.tipo == TipoLancamento.RECEITA)
                .where(Lancamento.data_lancamento >= first_day)
                .where(Lancamento.data_lancamento <= last_day)
            )
            lanc_result = await self.session.execute(lanc_query)
            valor_recebido_lanc = float(lanc_result.scalar_one())

            valor_recebido = valor_recebido_faturas + valor_recebido_lanc

            result.append({
                "mes": f"{year:04d}-{month:02d}",
                "valor_faturado": valor_faturado,
                "valor_recebido": valor_recebido,
            })

        return result

    async def gerar_relatorio_excel(
        self,
        data_inicio: date,
        data_fim: date,
        client_id: Optional[UUID] = None,
    ) -> io.BytesIO:
        """Generate financial report as Excel file."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()

        # Sheet 1: Faturas
        ws_faturas = wb.active
        ws_faturas.title = "Faturas"
        headers = ["Número", "Contrato", "Cliente", "Valor", "Valor Pago", "Vencimento", "Pagamento", "Status"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws_faturas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        query = (
            select(Fatura)
            .where(Fatura.tenant_id == self.tenant_id)
            .where(Fatura.data_vencimento >= data_inicio)
            .where(Fatura.data_vencimento <= data_fim)
            .order_by(Fatura.data_vencimento)
        )
        if client_id:
            query = query.where(Fatura.client_id == client_id)

        result = await self.session.execute(query)
        faturas = result.scalars().all()

        for row_num, fatura in enumerate(faturas, 2):
            ws_faturas.cell(row=row_num, column=1, value=fatura.numero)
            ws_faturas.cell(row=row_num, column=2, value=fatura.contrato.titulo if fatura.contrato else "")
            ws_faturas.cell(row=row_num, column=3, value=fatura.client.full_name if fatura.client else "")
            ws_faturas.cell(row=row_num, column=4, value=float(fatura.valor))
            ws_faturas.cell(row=row_num, column=5, value=float(fatura.valor_pago))
            ws_faturas.cell(row=row_num, column=6, value=fatura.data_vencimento.isoformat())
            ws_faturas.cell(row=row_num, column=7, value=fatura.data_pagamento.isoformat() if fatura.data_pagamento else "")
            ws_faturas.cell(row=row_num, column=8, value=fatura.status.value)

        # Auto-width columns
        for col in ws_faturas.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_faturas.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

        # Sheet 2: Lançamentos
        ws_lancamentos = wb.create_sheet("Lançamentos")
        l_headers = ["Tipo", "Categoria", "Descrição", "Valor", "Data", "Competência"]
        for col, header in enumerate(l_headers, 1):
            cell = ws_lancamentos.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        l_query = (
            select(Lancamento)
            .where(Lancamento.tenant_id == self.tenant_id)
            .where(Lancamento.data_lancamento >= data_inicio)
            .where(Lancamento.data_lancamento <= data_fim)
            .order_by(Lancamento.data_lancamento)
        )
        if client_id:
            l_query = l_query.where(Lancamento.client_id == client_id)

        l_result = await self.session.execute(l_query)
        lancamentos = l_result.scalars().all()

        for row_num, lanc in enumerate(lancamentos, 2):
            ws_lancamentos.cell(row=row_num, column=1, value=lanc.tipo.value)
            ws_lancamentos.cell(row=row_num, column=2, value=lanc.categoria.value)
            ws_lancamentos.cell(row=row_num, column=3, value=lanc.descricao)
            ws_lancamentos.cell(row=row_num, column=4, value=float(lanc.valor))
            ws_lancamentos.cell(row=row_num, column=5, value=lanc.data_lancamento.isoformat())
            ws_lancamentos.cell(row=row_num, column=6, value=lanc.data_competencia.isoformat() if lanc.data_competencia else "")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def gerar_relatorio_pdf(
        self,
        data_inicio: date,
        data_fim: date,
        client_id: Optional[UUID] = None,
    ) -> io.BytesIO:
        """Generate financial report as PDF file."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, spaceAfter=12)
        elements.append(Paragraph("Relatório Financeiro", title_style))
        elements.append(Paragraph(
            f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
            styles["Normal"],
        ))
        elements.append(Spacer(1, 10 * mm))

        # Revenue summary
        resumo = await self.fatura_repo.get_revenue_summary(data_inicio, data_fim)
        summary_data = [
            ["Métrica", "Valor (R$)"],
            ["Total Faturado", f"{resumo['total_faturado']:,.2f}"],
            ["Total Recebido", f"{resumo['total_recebido']:,.2f}"],
            ["A Receber", f"{resumo['total_a_receber']:,.2f}"],
            ["Em Atraso", f"{resumo['total_em_atraso']:,.2f}"],
        ]
        summary_table = Table(summary_data, colWidths=[200, 150])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 10 * mm))

        # Invoices table
        elements.append(Paragraph("Faturas", styles["Heading2"]))
        query = (
            select(Fatura)
            .where(Fatura.tenant_id == self.tenant_id)
            .where(Fatura.data_vencimento >= data_inicio)
            .where(Fatura.data_vencimento <= data_fim)
            .order_by(Fatura.data_vencimento)
        )
        if client_id:
            query = query.where(Fatura.client_id == client_id)

        result = await self.session.execute(query)
        faturas = result.scalars().all()

        if faturas:
            invoice_data = [["Número", "Valor", "Vencimento", "Status"]]
            for f in faturas:
                invoice_data.append([
                    f.numero,
                    f"R$ {float(f.valor):,.2f}",
                    f.data_vencimento.strftime("%d/%m/%Y"),
                    f.status.value.upper(),
                ])
            inv_table = Table(invoice_data, colWidths=[120, 100, 100, 80])
            inv_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ]))
            elements.append(inv_table)
        else:
            elements.append(Paragraph("Nenhuma fatura no período.", styles["Normal"]))

        doc.build(elements)
        output.seek(0)
        return output
